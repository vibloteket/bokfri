#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.11"
# dependencies = [
#   "openpyxl==3.1.5",
#   "xlrd==2.0.2",
#   "xlwt==1.3.0",
# ]
# ///
"""Generate reviewable Bokfri account-plan drafts from official BAS files.

Run from anywhere in the repository:

    uv run tools/account-plans/generate_account_plans.py

The script downloads checksum-pinned source files, reads the old Bokfri plans
for conservative VAT-code inheritance, writes old-style .xls files understood
by Bokfri, and creates Markdown/JSON review reports. It never modifies Bokfri's
packaged defaults unless --install is explicitly supplied.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import shutil
import sys
import urllib.request
from collections import defaultdict
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Iterable

import xlrd
import xlwt
from openpyxl import load_workbook

YEAR = 2026
ACCOUNT_PLAN_TYPE = "EUBAS97"

# Update this manifest once per BAS release. Checksums deliberately make the
# process fail if BAS replaces a file behind an unchanged URL.
SOURCES = {
    "accounts": {
        "filename": "BAS_kontoplan_2026_v2.xlsx",
        "url": "https://www.bas.se/wp-content/uploads/2026/04/BAS_kontoplan_2026_v2.xlsx",
        "sha256": "a86b39937fab280d4e5db895c04c2af6e145695863d4845ff14eea5d0302328a",
        "published_for": "BAS 2026",
    },
    "ink2": {
        "filename": "INK2_P1_intervall-241119.xlsx",
        "url": "https://www.bas.se/wp-content/uploads/2024/11/INK2_P1_intervall-241119.xlsx",
        "sha256": "59bc463c3e5f75457cd16917c9e9b3fcc25418ea232d9c0c5955d0a72a8638ee",
        "published_for": "BAS 2023 / INK2",
    },
    "ink3": {
        "filename": "INK3_P1_Intervall-241119.xlsx",
        "url": "https://www.bas.se/wp-content/uploads/2024/11/INK3_P1_Intervall-241119.xlsx",
        "sha256": "414d153dd0f5812cfd6864462e3e2223f6bbaba12b648a023183f79a9b6711c5",
        "published_for": "BAS 2024 / INK3",
    },
    "ink4": {
        "filename": "INK4_P1_Intervall-241119.xlsx",
        "url": "https://www.bas.se/wp-content/uploads/2024/11/INK4_P1_Intervall-241119.xlsx",
        "sha256": "5130ddaf7751be690236498a1119242bb1942fc105c19bf0cac650bb184a13aa",
        "published_for": "BAS 2024 / INK4",
    },
    "ne": {
        "filename": "NE_EJ_K1-Intervall-231002.xlsx",
        "url": "https://www.bas.se/wp-content/uploads/2023/10/NE_EJ_K1-Intervall-231002.xlsx",
        "sha256": "25fe5f804299699bf3c03a4145b8a76bbe386cc08b71e572858afde1d6ec9fb5",
        "published_for": "BAS 2023 / NE ej K1",
    },
    "ne_k1": {
        "filename": "NE_K1-201002.xlsx",
        "url": "https://www.bas.se/wp-content/uploads/2023/10/NE_K1-201002.xlsx",
        "sha256": "15234aa24f1f127d8791fc8ce1d1a30b80c2d6503c44c19a488a50213562dc34",
        "published_for": "BAS Förenklat årsbokslut 2023 / NE K1",
    },
}

PLAN_SPECS = {
    "ink2_ab": "BAS 2026 - Aktiebolag",
    "ink2_ef": "BAS 2026 - Ekonomisk förening",
    "ink3": "BAS 2026 - Ideell förening, stiftelse och trossamfund",
    "ink4": "BAS 2026 - Handelsbolag och kommanditbolag",
    "ne": "BAS 2026 - Enskild firma, ej K1",
    "ne_k1": "BAS 2026 - Enskild firma K1",
}

PLAN_SRU_SOURCE = {
    "ink2_ab": "ink2",
    "ink2_ef": "ink2",
    "ink3": "ink3",
    "ink4": "ink4",
    "ne": "ne",
}

# BAS deliberately assigns account 2087 two meanings for different legal forms.
# Keeping it in one combined plan would silently attach the wrong name to one of
# them. Other plan types omit this organization-specific account.
ACCOUNT_NAME_VARIANTS = {
    "ink2_ab": {2087: "Bunden överkursfond"},
    "ink2_ef": {2087: "Insatsemission"},
}

# Reviewed decisions for old VAT mappings that conflict between Bokfri plans.
# Blank means that the modern BAS account is too broad for one automatic VAT
# treatment. The reduced K1 account 3100 is explicitly named Momsfria intäkter.
SETTLEMENT_VAT_OVERRIDES = {
    # Old BAS plans used 1480/2480 as VAT settlement accounts. Those account
    # numbers have different meanings in BAS 2026; only the modern settlement
    # accounts may carry Bokfri's R1/R2 markers.
    1480: "",
    1650: "R1",
    2480: "",
    2650: "R2",
    3740: "A",
}

VAT_OVERRIDES = {
    "ink2_ab": {**SETTLEMENT_VAT_OVERRIDES, 3100: "", 3630: "", 3910: "", 3920: ""},
    "ink2_ef": {**SETTLEMENT_VAT_OVERRIDES, 3100: "", 3630: "", 3910: "", 3920: ""},
    "ink3": {**SETTLEMENT_VAT_OVERRIDES, 3100: "", 3630: "", 3910: "", 3920: ""},
    "ink4": {**SETTLEMENT_VAT_OVERRIDES, 3100: "", 3630: "", 3910: "", 3920: ""},
    "ne": {**SETTLEMENT_VAT_OVERRIDES, 3100: "", 3630: "", 3910: "", 3920: ""},
    "ne_k1": {**SETTLEMENT_VAT_OVERRIDES, 3100: "MF"},
}


@dataclass
class Account:
    number: int
    name: str
    vat_code: str = ""
    sru_code: str = ""
    report_code: str = ""
    not_k2: bool = False


@dataclass
class Issue:
    severity: str
    category: str
    message: str


def repository_root() -> Path:
    path = Path(__file__).resolve()
    for parent in path.parents:
        if (parent / "pom.xml").exists() and (parent / "src/main/resources").exists():
            return parent
    raise RuntimeError("Could not locate the Bokfri repository root")


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def obtain_sources(cache: Path, offline: bool) -> dict[str, Path]:
    cache.mkdir(parents=True, exist_ok=True)
    paths: dict[str, Path] = {}
    for key, source in SOURCES.items():
        path = cache / source["filename"]
        if not path.exists():
            if offline:
                raise RuntimeError(f"Missing cached source in offline mode: {path}")
            print(f"Downloading {source['url']}")
            request = urllib.request.Request(source["url"], headers={"User-Agent": "Bokfri account-plan generator"})
            with urllib.request.urlopen(request, timeout=60) as response, path.open("wb") as output:
                shutil.copyfileobj(response, output)
        actual = sha256(path)
        if actual != source["sha256"]:
            raise RuntimeError(
                f"Checksum mismatch for {path.name}:\n"
                f" expected {source['sha256']}\n actual   {actual}\n"
                "Review the upstream change and update SOURCES intentionally."
            )
        paths[key] = path
    return paths


def cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return str(int(value)) if float(value).is_integer() else str(value)
    return str(value).strip()


def account_number(value: object) -> tuple[int, bool] | None:
    text = cell_text(value)
    match = re.match(r"^(\d{4})(#?)$", text)
    if not match:
        return None
    return int(match.group(1)), bool(match.group(2))


def read_bas_account_variants(path: Path, issues: list[Issue]) -> dict[int, list[Account]]:
    sheet = load_workbook(path, data_only=False, read_only=True).active
    variants: dict[int, list[Account]] = defaultdict(list)
    for row in sheet.iter_rows(values_only=True):
        for number_col, name_col in ((0, 1), (2, 3)):
            parsed = account_number(row[number_col] if len(row) > number_col else None)
            if parsed is None:
                continue
            number, not_k2 = parsed
            name = cell_text(row[name_col] if len(row) > name_col else None).replace("\n", " ")
            if not name:
                issues.append(Issue("error", "accounts", f"Account {number} has no name"))
                continue
            existing = next((account for account in variants[number] if account.name == name), None)
            if existing:
                existing.not_k2 |= not_k2
            else:
                variants[number].append(Account(number, name, not_k2=not_k2))
    return variants


def select_bas_accounts(
    variants: dict[int, list[Account]], plan: str, issues: list[Issue]
) -> dict[int, Account]:
    accounts: dict[int, Account] = {}
    plan_variants = ACCOUNT_NAME_VARIANTS.get(plan, {})
    for number, choices in variants.items():
        if len(choices) == 1:
            selected = choices[0]
        elif number in plan_variants:
            wanted = plan_variants[number]
            selected = next((account for account in choices if account.name == wanted), None)
            if selected is None:
                issues.append(Issue("error", "account-variant", f"{plan}: missing expected {number} {wanted!r}"))
                continue
        else:
            names = " / ".join(repr(account.name) for account in choices)
            issues.append(
                Issue(
                    "info",
                    "account-variant",
                    f"{plan}: omitted organization-specific account {number}: {names}",
                )
            )
            continue
        accounts[number] = Account(**asdict(selected))
    return accounts


def read_k1_accounts(path: Path, issues: list[Issue]) -> dict[int, Account]:
    """K1 uses BAS's deliberately reduced, aggregate account plan."""
    sheet = load_workbook(path, data_only=True, read_only=True).active
    accounts: dict[int, Account] = {}
    for row in sheet.iter_rows(values_only=True):
        field_code = cell_text(row[0] if len(row) > 0 else None)
        report_code = cell_text(row[1] if len(row) > 1 else None).strip()
        number_text = cell_text(row[3] if len(row) > 3 else None)
        name = cell_text(row[4] if len(row) > 4 else None).replace("\n", " ")
        if not (field_code.isdigit() and re.fullmatch(r"\d{4}", number_text) and name):
            continue
        number = int(number_text)
        if number in accounts:
            issues.append(Issue("error", "duplicate-account", f"Duplicate K1 account {number}"))
            continue
        accounts[number] = Account(number, name, sru_code=field_code, report_code=report_code)
    return accounts


def expand_token(token: str) -> set[int]:
    """Expand one BAS expression token such as 112x, 84xx or 1000-1087."""
    token = token.strip().replace("–", "-").replace("−", "-")
    token = re.sub(r"\(.*?\)", "", token).strip()
    token = token.lstrip("+- ")
    if re.fullmatch(r"\d{4}", token):
        return {int(token)}
    if re.fullmatch(r"\d{1,3}x{1,3}", token, re.IGNORECASE) and len(token) == 4:
        start = int(re.sub("[xX]", "0", token))
        end = int(re.sub("[xX]", "9", token))
        return set(range(start, end + 1))
    match = re.fullmatch(r"(\d{4})\s*-\s*(\d{4})", token)
    if match:
        start, end = map(int, match.groups())
        return set(range(min(start, end), max(start, end) + 1))
    match = re.fullmatch(r"(\d+)(x+)\s*-\s*(\d+)(x+)", token, re.IGNORECASE)
    if match:
        left, left_x, right, right_x = match.groups()
        if len(left + left_x) == len(right + right_x) == 4:
            start = int(left + "0" * len(left_x))
            end = int(right + "9" * len(right_x))
            return set(range(min(start, end), max(start, end) + 1))
    return set()


def expand_expression(expression: str) -> tuple[set[int], list[str]]:
    numbers: set[int] = set()
    unknown: list[str] = []
    for token in re.split(r"[,;]", expression.replace("\n", ",")):
        token = token.strip()
        if not token:
            continue
        expanded = expand_token(token)
        if expanded:
            numbers.update(expanded)
        else:
            unknown.append(token)
    return numbers, unknown


def read_sru_candidates(path: Path, valid_accounts: set[int], issues: list[Issue]) -> dict[int, set[str]]:
    sheet = load_workbook(path, data_only=True, read_only=True).active
    candidates: dict[int, set[str]] = defaultdict(set)
    for row in sheet.iter_rows(values_only=True):
        field_code = cell_text(row[0] if len(row) > 0 else None)
        expression = cell_text(row[3] if len(row) > 3 else None)
        if not (re.fullmatch(r"\d{4}", field_code) and expression):
            continue
        numbers, unknown = expand_expression(expression)
        for token in unknown:
            # BAS uses prose/footnotes where a declaration field intentionally
            # has no account mapping. Preserve this in the report as info.
            issues.append(Issue("info", "sru-note", f"{path.name}: {field_code}: {token}"))
        for number in numbers & valid_accounts:
            candidates[number].add(field_code)
    return candidates


def apply_sru(accounts: dict[int, Account], candidates: dict[int, set[str]], issues: list[Issue], plan: str) -> None:
    ambiguous: dict[tuple[str, ...], list[int]] = defaultdict(list)
    for number, codes in candidates.items():
        if len(codes) == 1:
            accounts[number].sru_code = next(iter(codes))
        elif len(codes) > 1:
            ambiguous[tuple(sorted(codes))].append(number)
    for codes, numbers in sorted(ambiguous.items()):
        preview = ", ".join(map(str, sorted(numbers)[:12]))
        if len(numbers) > 12:
            preview += f", … ({len(numbers)} accounts)"
        issues.append(
            Issue(
                "warning",
                "ambiguous-sru",
                f"{plan}: {preview} match {'/'.join(codes)}; left blank because Bokfri stores one code",
            )
        )


def read_old_vat_codes(defaults: Path, issues: list[Issue]) -> tuple[dict[int, str], dict[int, list[str]]]:
    candidates: dict[int, set[str]] = defaultdict(set)
    for path in sorted(defaults.glob("*.xls")):
        book = xlrd.open_workbook(path)
        sheet = book.sheet_by_index(0)
        for row in range(5, sheet.nrows):
            number_text = str(sheet.cell_value(row, 0)).strip().removesuffix(".0")
            vat_code = str(sheet.cell_value(row, 2)).strip()
            if re.fullmatch(r"\d{4}", number_text) and vat_code:
                candidates[int(number_text)].add(vat_code)
    consensus = {number: next(iter(codes)) for number, codes in candidates.items() if len(codes) == 1}
    conflicts = {number: sorted(codes) for number, codes in candidates.items() if len(codes) > 1}
    return consensus, conflicts


def apply_vat_codes(accounts: dict[int, Account], vat_codes: dict[int, str], plan: str) -> int:
    applied = 0
    overrides = VAT_OVERRIDES.get(plan, {})
    for number, account in accounts.items():
        if number in overrides:
            account.vat_code = overrides[number]
        elif number in vat_codes:
            account.vat_code = vat_codes[number]
        if account.vat_code:
            applied += 1
    return applied


def write_xls(path: Path, name: str, accounts: Iterable[Account]) -> None:
    workbook = xlwt.Workbook(encoding="windows-1252")
    sheet = workbook.add_sheet(name[:31])
    metadata = (("#Namn", name), ("#Typ", ACCOUNT_PLAN_TYPE), ("#Taxeringsår", str(YEAR)), ("#Start", 6))
    for row, values in enumerate(metadata):
        for column, value in enumerate(values):
            sheet.write(row, column, value)
    for column, heading in enumerate(("Konto nr", "Beskrivning", "Momskod", "SRU-kod", "Rapportkod")):
        sheet.write(4, column, heading)
    for row, account in enumerate(sorted(accounts, key=lambda item: item.number), start=5):
        for column, value in enumerate(
            (account.number, account.name, account.vat_code, account.sru_code, account.report_code)
        ):
            sheet.write(row, column, value)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(path))


def verify_xls(path: Path, expected: dict[int, Account]) -> None:
    sheet = xlrd.open_workbook(path).sheet_by_index(0)
    actual: dict[int, tuple[str, str, str, str]] = {}
    for row in range(5, sheet.nrows):
        number = int(sheet.cell_value(row, 0))
        actual[number] = tuple(str(sheet.cell_value(row, col)).strip() for col in range(1, 5))
    if set(actual) != set(expected):
        raise RuntimeError(f"Round-trip account mismatch in {path}")
    for number, account in expected.items():
        wanted = (account.name, account.vat_code, account.sru_code, account.report_code)
        if actual[number] != wanted:
            raise RuntimeError(f"Round-trip value mismatch for account {number} in {path}")


def safe_filename(name: str) -> str:
    replacements = str.maketrans({"å": "a", "ä": "a", "ö": "o", "Å": "A", "Ä": "A", "Ö": "O"})
    return re.sub(r"[^A-Za-z0-9_-]+", "-", name.translate(replacements)).strip("-") + ".xls"


def write_reports(
    output: Path,
    plans: dict[str, dict[int, Account]],
    issues: list[Issue],
    vat_applied: dict[str, int],
) -> None:
    report = {
        "year": YEAR,
        "sources": SOURCES,
        "plans": {
            key: {
                "name": PLAN_SPECS[key],
                "accounts": len(accounts),
                "with_vat_code": sum(bool(a.vat_code) for a in accounts.values()),
                "with_sru_code": sum(bool(a.sru_code) for a in accounts.values()),
                "with_report_code": sum(bool(a.report_code) for a in accounts.values()),
                "not_k2": sum(a.not_k2 for a in accounts.values()),
                "vat_codes_inherited": vat_applied[key],
            }
            for key, accounts in plans.items()
        },
        "issues": [asdict(issue) for issue in issues],
    }
    (output / "review.json").write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    lines = [
        f"# Bokfri account-plan generation review — {YEAR}",
        "",
        "Generated from checksum-pinned official BAS files. This is a review artifact, not proof that tax/VAT mappings are correct.",
        "",
        "## Plans",
        "",
        "| Plan | Accounts | VAT | SRU | Report codes | Not K2 |",
        "|---|---:|---:|---:|---:|---:|",
    ]
    for key, data in report["plans"].items():
        lines.append(
            f"| {data['name']} | {data['accounts']} | {data['with_vat_code']} | "
            f"{data['with_sru_code']} | {data['with_report_code']} | {data['not_k2']} |"
        )
    lines += [
        "",
        "## Important limitations",
        "",
        "- The free BAS account workbook has account numbers/names but no VAT or SRU mappings.",
        "- VAT codes are inherited only when all old Bokfri plans agree for an account number; they still require manual review.",
        "- BAS SRU interval files target BAS 2023/2024 while account names come from BAS 2026.",
        "- Account 2087 has different official meanings for AB and EF, so those are separate plans; other plan types omit it.",
        "- Sign-dependent SRU mappings cannot be represented by Bokfri's current single SRU field and are left blank when ambiguous.",
        "- The K1 plan comes from BAS's reduced NE K1 account table and includes B/R report codes.",
        "",
        "## Issues",
        "",
    ]
    if not issues:
        lines.append("No issues found.")
    else:
        for issue in issues:
            lines.append(f"- **{issue.severity.upper()} / {issue.category}:** {issue.message}")
    (output / "review.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def self_test() -> None:
    cases = {
        "112x": {1120, 1129},
        "84xx": {8400, 8499},
        "1000-1002": {1000, 1001, 1002},
        "151x–152x": {1510, 1529},
        "30xx-37xx": {3000, 3799},
        "+ 899x": {8990, 8999},
        "8000-8001 (Om netto +)": {8000, 8001},
    }
    for expression, bounds in cases.items():
        values = expand_token(expression)
        low, high = min(bounds), max(bounds)
        assert min(values) == low and max(values) == high, (expression, min(values), max(values))
    assert expand_expression("112x, 1200-1201")[1] == []


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--offline", action="store_true", help="Use cache only; do not download")
    parser.add_argument("--install", action="store_true", help="Replace packaged default plans after generation")
    parser.add_argument("--output", type=Path, help="Output directory (default: tools/account-plans/generated/2026)")
    args = parser.parse_args()

    self_test()
    root = repository_root()
    work_dir = root / "tools/account-plans"
    cache = work_dir / "cache"
    output = args.output.resolve() if args.output else work_dir / "generated" / str(YEAR)
    output.mkdir(parents=True, exist_ok=True)

    issues: list[Issue] = []
    source_paths = obtain_sources(cache, args.offline)
    bas_variants = read_bas_account_variants(source_paths["accounts"], issues)
    vat_codes, vat_conflicts = read_old_vat_codes(root / "src/main/resources/account/default", issues)
    for number, codes in vat_conflicts.items():
        issues.append(
            Issue(
                "info",
                "reviewed-vat",
                f"Account {number}: old plans disagree ({', '.join(codes)}); resolved by explicit per-plan override",
            )
        )

    plans: dict[str, dict[int, Account]] = {}
    vat_applied: dict[str, int] = {}
    for key, sru_source in PLAN_SRU_SOURCE.items():
        accounts = select_bas_accounts(bas_variants, key, issues)
        candidates = read_sru_candidates(source_paths[sru_source], set(accounts), issues)
        apply_sru(accounts, candidates, issues, key)
        vat_applied[key] = apply_vat_codes(accounts, vat_codes, key)
        plans[key] = accounts

    k1_accounts = read_k1_accounts(source_paths["ne_k1"], issues)
    vat_applied["ne_k1"] = apply_vat_codes(k1_accounts, vat_codes, "ne_k1")
    plans["ne_k1"] = k1_accounts

    generated_files: dict[str, Path] = {}
    for key, accounts in plans.items():
        path = output / safe_filename(PLAN_SPECS[key])
        write_xls(path, PLAN_SPECS[key], accounts.values())
        verify_xls(path, accounts)
        generated_files[key] = path

    write_reports(output, plans, issues, vat_applied)

    if args.install:
        errors = [issue for issue in issues if issue.severity == "error"]
        if errors:
            print(f"Refusing --install: {len(errors)} error(s) remain; see {output / 'review.md'}", file=sys.stderr)
            return 2
        defaults = root / "src/main/resources/account/default"
        for old_file in defaults.glob("*.xls"):
            old_file.unlink()
        for path in generated_files.values():
            shutil.copy2(path, defaults / path.name)
        print(f"Installed {len(generated_files)} account plans in {defaults}")

    counts = {
        severity: sum(issue.severity == severity for issue in issues)
        for severity in ("error", "warning", "info")
    }
    print(f"Generated {len(plans)} plans in {output}")
    print(
        f"Review: {output / 'review.md'} "
        f"({counts['error']} errors, {counts['warning']} warnings, {counts['info']} info)"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
